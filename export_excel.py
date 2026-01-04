"""
Excel导出模块
提供按班级和按老师导出课表为Excel文件的功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


class ExcelExporter:
    """课表Excel导出器"""
    
    def __init__(self):
        self.days = ["周一", "周二", "周三", "周四", "周五"]
        self.periods = 9
        
        # 样式定义
        self.header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.sub_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        self.empty_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _apply_cell_style(self, cell, is_header=False, is_sub=False, is_empty=False):
        """应用单元格样式"""
        cell.border = self.border
        cell.alignment = self.center_align
        
        if is_header:
            cell.fill = self.header_fill
            cell.font = self.header_font
        elif is_sub:
            cell.fill = self.sub_fill
        elif is_empty:
            cell.fill = self.empty_fill
    
    def export_class_schedule(self, schedule_data, class_id):
        """导出单个班级的课表
        
        Args:
            schedule_data: 完整课表数据字典
            class_id: 班级ID（如 "1"）
            
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"{class_id}班课表"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
        
        # 标题行
        ws['A1'] = '节次'
        for col_idx, day in enumerate(self.days, start=2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = day
            self._apply_cell_style(cell, is_header=True)
        
        self._apply_cell_style(ws['A1'], is_header=True)
        
        # 获取班级课表数据 - class_id可能是整数或字符串，都尝试
        class_schedule = schedule_data.get(int(class_id), schedule_data.get(str(class_id), {}))
        
        # 填充数据
        for period in range(self.periods):
            row_idx = period + 2
            period_cell = ws.cell(row=row_idx, column=1)
            period_cell.value = f"第{period + 1}节"
            self._apply_cell_style(period_cell, is_header=True)
            
            for day in range(5):
                col_idx = day + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 获取课程信息 - 使用整数键
                period_data = class_schedule.get(period, {})
                info = period_data.get(day)
                
                if info:
                    # 格式：科目名\n老师名
                    cell.value = f"{info['subject']}\n{info['teacher_name']}"
                    
                    # 根据状态设置样式
                    is_sub = info.get('is_sub', False)
                    is_empty = (info['teacher_name'] == "【自习】")
                    
                    self._apply_cell_style(cell, is_sub=is_sub, is_empty=is_empty)
                    
                    # 设置行高以适应两行文字
                    ws.row_dimensions[row_idx].height = 40
                else:
                    cell.value = ""
                    self._apply_cell_style(cell)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_all_classes(self, schedule_data):
        """导出所有班级的课表（多sheet）
        
        Args:
            schedule_data: 完整课表数据字典
            
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet
        
        for class_id in sorted(schedule_data.keys(), key=lambda x: int(x)):
            ws = wb.create_sheet(title=f"{class_id}班")
            
            # 设置列宽
            ws.column_dimensions['A'].width = 10
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 18
            
            # 标题行
            ws['A1'] = '节次'
            for col_idx, day in enumerate(self.days, start=2):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = day
                self._apply_cell_style(cell, is_header=True)
            
            self._apply_cell_style(ws['A1'], is_header=True)
            
            # 获取班级课表数据 - class_id可能是整数或字符串
            class_schedule = schedule_data.get(int(class_id), schedule_data.get(str(class_id), {}))
            
            # 填充数据
            for period in range(self.periods):
                row_idx = period + 2
                period_cell = ws.cell(row=row_idx, column=1)
                period_cell.value = f"第{period + 1}节"
                self._apply_cell_style(period_cell, is_header=True)
                
                for day in range(5):
                    col_idx = day + 2
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # 获取课程信息 - 使用整数键
                    period_data = class_schedule.get(period, {})
                    info = period_data.get(day)
                    
                    if info:
                        cell.value = f"{info['subject']}\n{info['teacher_name']}"
                        
                        is_sub = info.get('is_sub', False)
                        is_empty = (info['teacher_name'] == "【自习】")
                        
                        self._apply_cell_style(cell, is_sub=is_sub, is_empty=is_empty)
                        ws.row_dimensions[row_idx].height = 40
                    else:
                        cell.value = ""
                        self._apply_cell_style(cell)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_teacher_schedule(self, schedule_data, teachers_db, teacher_name):
        """导出单个老师的课表
        
        Args:
            schedule_data: 完整课表数据字典
            teachers_db: 老师数据库列表
            teacher_name: 老师姓名
            
        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"{teacher_name}的课表"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 10
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
        
        # 标题行
        ws['A1'] = '节次'
        for col_idx, day in enumerate(self.days, start=2):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = day
            self._apply_cell_style(cell, is_header=True)
        
        self._apply_cell_style(ws['A1'], is_header=True)
        
        # 构建老师的课表矩阵
        teacher_schedule = [[None for _ in range(5)] for _ in range(self.periods)]
        
        # 遍历所有班级的所有课程
        for class_id, class_data in schedule_data.items():
            for period_str, period_data in class_data.items():
                period = int(period_str)
                for day_str, info in period_data.items():
                    day = int(day_str)
                    if info and info['teacher_name'] == teacher_name:
                        teacher_schedule[period][day] = {
                            'class_id': class_id,
                            'subject': info['subject'],
                            'is_sub': info.get('is_sub', False)
                        }
        
        # 填充数据
        for period in range(self.periods):
            row_idx = period + 2
            period_cell = ws.cell(row=row_idx, column=1)
            period_cell.value = f"第{period + 1}节"
            self._apply_cell_style(period_cell, is_header=True)
            
            for day in range(5):
                col_idx = day + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                
                info = teacher_schedule[period][day]
                
                if info:
                    # 格式：班级\n科目
                    cell.value = f"{info['class_id']}班\n{info['subject']}"
                    
                    is_sub = info.get('is_sub', False)
                    self._apply_cell_style(cell, is_sub=is_sub)
                    ws.row_dimensions[row_idx].height = 40
                else:
                    cell.value = ""
                    self._apply_cell_style(cell)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
